"""Exports Accounts + Businesses + BusinessQuestionnaires (joined) to an .xlsx file for review.

Usage:
    export DATABASE_URI=postgresql://user:pass@host:port/db
    python scripts/export_business_questionnaires.py [output_path]
"""

import os
import sys
from pathlib import Path

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Mirrors the int-backed enums in backend/src/Vira.Abstractions/Common/*.cs — keep these lists
# in sync (same order) if those enums ever change.
CREATOR_CATEGORY = [
    "Food", "Sport", "Tech", "Beauty", "Travel",
    "Comedy", "Education", "Lifestyle", "Gaming", "Music",
]
COMPANY_SIZE = ["Solo", "Small", "Medium", "Large"]
BUDGET_BAND = ["Under1k", "From1kTo5k", "From5kTo20k", "Over20k"]
AUDIENCE_AGE = ["Teens", "A18_24", "A25_34", "A35_44", "A45Plus"]
CAMPAIGN_OBJECTIVE = ["Awareness", "Visits", "Offer", "Launch", "Community"]

OUTPUT_XLSX = Path(__file__).resolve().parent / "business_questionnaires.xlsx"

QUERY = """
SELECT
    a."Email" AS email,
    a."Type" AS account_type,
    a."FirebaseUid" AS firebase_uid,
    b."Id" AS business_id,
    b."CompanyName" AS company_name,
    b."CreatedAt" AS business_created_at,
    q."Verticals" AS verticals,
    q."CompanySize" AS company_size,
    q."BudgetBand" AS budget_band,
    q."TargetAudienceAges" AS target_audience_ages,
    q."PrimaryGoal" AS primary_goal,
    q."AvoidsAlcohol" AS avoids_alcohol,
    q."AvoidsGambling" AS avoids_gambling,
    q."AvoidsPolitical" AS avoids_political,
    q."Description" AS description,
    q."Values" AS brand_values,
    q."Website" AS website,
    q."CompetitorBrands" AS competitor_brands,
    q."ProductsToPromote" AS products_to_promote,
    q."CreatedAt" AS questionnaire_created_at
FROM "Businesses" b
JOIN "Accounts" a ON a."Id" = b."AccountId"
LEFT JOIN "BusinessQuestionnaires" q ON q."BusinessId" = b."Id"
ORDER BY a."Email";
"""

COLUMNS = [
    "Email", "AccountType", "HasFirebaseAuth",
    "BusinessId", "CompanyName", "BusinessCreatedAt",
    "Verticals", "CompanySize", "BudgetBand", "TargetAudienceAges", "PrimaryGoal",
    "AvoidsAlcohol", "AvoidsGambling", "AvoidsPolitical",
    "Description", "Values", "Website", "CompetitorBrands", "ProductsToPromote",
    "QuestionnaireCreatedAt",
]


class BusinessQuestionnaireExporter:
    """Joins Accounts + Businesses + BusinessQuestionnaires and writes the result to .xlsx,
    decoding the int-backed enum/array columns into readable labels."""

    def __init__(self, dsn: str | None = None) -> None:
        self.dsn = dsn or os.environ["DATABASE_URI"]

    def fetch_rows(self) -> list[dict]:
        with psycopg2.connect(self.dsn) as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(QUERY)
                return [dict(row) for row in cur.fetchall()]

    def export(self, output_path: Path = OUTPUT_XLSX) -> Path:
        rows = self.fetch_rows()

        wb = Workbook()
        ws = wb.active
        ws.title = "BusinessQuestionnaires"
        ws.append(COLUMNS)

        for row in rows:
            ws.append([
                row["email"],
                "Business" if row["account_type"] == 1 else "Creator",
                row["firebase_uid"] is not None,
                str(row["business_id"]),
                row["company_name"],
                self._fmt_dt(row["business_created_at"]),
                self._decode_list(CREATOR_CATEGORY, row["verticals"]),
                self._decode(COMPANY_SIZE, row["company_size"]),
                self._decode(BUDGET_BAND, row["budget_band"]),
                self._decode_list(AUDIENCE_AGE, row["target_audience_ages"]),
                self._decode(CAMPAIGN_OBJECTIVE, row["primary_goal"]),
                row["avoids_alcohol"],
                row["avoids_gambling"],
                row["avoids_political"],
                row["description"],
                self._join(row["brand_values"]),
                row["website"],
                self._join(row["competitor_brands"]),
                row["products_to_promote"],
                self._fmt_dt(row["questionnaire_created_at"]),
            ])

        self._autofit(ws)
        wb.save(output_path)
        return output_path

    @staticmethod
    def _decode(labels: list[str], value: int | None) -> str | None:
        if value is None:
            return None
        return labels[value] if 0 <= value < len(labels) else f"UNKNOWN({value})"

    @classmethod
    def _decode_list(cls, labels: list[str], values: list[int] | None) -> str | None:
        if values is None:
            return None
        return ", ".join(cls._decode(labels, v) for v in values)

    @staticmethod
    def _join(values: list[str] | None) -> str | None:
        if values is None:
            return None
        return ", ".join(values)

    @staticmethod
    def _fmt_dt(dt) -> str | None:
        return dt.isoformat() if dt else None

    @staticmethod
    def _autofit(ws, padding: int = 2, max_width: int = 60) -> None:
        for col_cells in ws.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + padding, max_width)


def main() -> None:
    output_path = Path(sys.argv[1]) if len(sys.argv) > 1 else OUTPUT_XLSX
    path = BusinessQuestionnaireExporter().export(output_path)
    print(f"saved to {path}")


if __name__ == "__main__":
    main()
